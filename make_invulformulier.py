#!/usr/bin/env python3
"""Tempetoto 2026 — Excel invulformulier (één template, naam bovenaan invullen)"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ─── Data (identiek aan data.js) ─────────────────────────────────────────────

GROUPS = {
    'A': ["Mexico","Zuid-Afrika","Zuid-Korea","Tsjechië"],
    'B': ["Canada","Zwitserland","Qatar","Bosnië-Herzegovina"],
    'C': ["Brazilië","Marokko","Haïti","Schotland"],
    'D': ["Verenigde Staten","Paraguay","Australië","Turkije"],
    'E': ["Duitsland","Curaçao","Ivoorkust","Ecuador"],
    'F': ["Nederland","Japan","Zweden","Tunesië"],
    'G': ["België","Egypte","Iran","Nieuw-Zeeland"],
    'H': ["Spanje","Kaapverdië","Saoedi-Arabië","Uruguay"],
    'I': ["Frankrijk","Senegal","Noorwegen","Irak"],
    'J': ["Argentinië","Algerije","Oostenrijk","Jordanië"],
    'K': ["Portugal","DR Congo","Oezbekistan","Colombia"],
    'L': ["Engeland","Kroatië","Ghana","Panama"],
}

TEAMCOLORS = {
    "Mexico":"006847","Zuid-Afrika":"ffb612","Zuid-Korea":"cd2e3a","Tsjechië":"11457e",
    "Canada":"ff0000","Zwitserland":"000000","Qatar":"8a1538","Bosnië-Herzegovina":"002395",
    "Brazilië":"009c3b","Marokko":"c1272d","Haïti":"00209f","Schotland":"5fb0e5",
    "Verenigde Staten":"3c3b6e","Paraguay":"d52b1e","Australië":"ffcd00","Turkije":"111111",
    "Duitsland":"000000","Curaçao":"002b7f","Ivoorkust":"f77f00","Ecuador":"ed1c24",
    "Nederland":"ff6a13","Japan":"bc002d","Zweden":"006aa7","Tunesië":"111111",
    "België":"fdda24","Egypte":"ce1126","Iran":"239f40","Nieuw-Zeeland":"00247d",
    "Spanje":"aa151b","Kaapverdië":"003893","Saoedi-Arabië":"006c35","Uruguay":"5b92e5",
    "Frankrijk":"0055a4","Senegal":"00853f","Noorwegen":"ba0c2f","Irak":"111111",
    "Argentinië":"74acdf","Algerije":"006233","Oostenrijk":"ed2939","Jordanië":"000000",
    "Portugal":"006600","DR Congo":"00aaff","Oezbekistan":"1eb53a","Colombia":"fcd116",
    "Engeland":"ce1124","Kroatië":"1742a0","Ghana":"ffd100","Panama":"111111",
}

RANKING = {
    "Frankrijk":1,"Spanje":2,"Argentinië":3,"Engeland":4,"Portugal":5,"Brazilië":6,
    "Nederland":7,"Marokko":8,"België":9,"Duitsland":10,"Kroatië":11,"Colombia":13,
    "Senegal":14,"Mexico":15,"Verenigde Staten":16,"Uruguay":17,"Japan":18,"Zwitserland":19,
    "Iran":21,"Oostenrijk":22,"Zuid-Korea":23,"Australië":24,"Ecuador":25,"Noorwegen":26,
    "Panama":30,"Egypte":32,"Algerije":36,"Paraguay":38,"Ivoorkust":40,"Tunesië":41,
    "Schotland":42,"Zweden":43,"Tsjechië":44,"Qatar":51,"Oezbekistan":54,"DR Congo":56,
    "Saoedi-Arabië":58,"Irak":59,"Zuid-Afrika":60,"Jordanië":64,"Kaapverdië":70,"Ghana":73,
    "Bosnië-Herzegovina":74,"Haïti":83,"Nieuw-Zeeland":86,"Curaçao":90,
}

FAVORIETEN = ["Frankrijk","Spanje","Argentinië","Engeland","Portugal","Brazilië",
              "Nederland","Marokko","België","Duitsland","Kroatië","Colombia"]

KO_ROUNDS = [
    {'key':'R32','naam':'Ronde van 32', 'toto':5, 'exact':3},
    {'key':'R16','naam':'Ronde van 16', 'toto':7, 'exact':4},
    {'key':'KF', 'naam':'Kwartfinales', 'toto':9, 'exact':5},
    {'key':'HF', 'naam':'Halve finales','toto':11,'exact':6},
    {'key':'F',  'naam':'Finale',       'toto':13,'exact':7},
]

KO_BRACKETS = {
    'R32': [
        ("2A","2B"),("1E","3e (A/B/C/D/F)"),("1F","2C"),("1C","2F"),
        ("1I","3e (C/D/F/G/H)"),("2E","2I"),("1A","3e (C/E/F/H/I)"),("1L","3e (E/H/I/J/K)"),
        ("1D","3e (B/E/F/I/J)"),("1G","3e (A/E/H/I/J)"),("2K","2L"),("1H","2J"),
        ("1B","3e (E/F/G/I/J)"),("1J","2H"),("1K","3e (D/E/I/J/L)"),("2D","2G"),
    ],
    'R16': [
        ("W32-2","W32-5"),("W32-1","W32-3"),("W32-4","W32-6"),("W32-7","W32-8"),
        ("W32-11","W32-12"),("W32-9","W32-10"),("W32-14","W32-16"),("W32-13","W32-15"),
    ],
    'KF': [("W16-1","W16-2"),("W16-5","W16-6"),("W16-3","W16-4"),("W16-7","W16-8")],
    'HF': [("WKF-1","WKF-2"),("WKF-3","WKF-4")],
    'F':  [("WHF-1","WHF-2")],
}

SPELREGELS = [
    "Vul ALLE 72 groepswedstrijden in als '2-1' (thuis–uit, uitslag na 90 min). Dit zijn de lichtblauwe vakjes.",
    "Vul ook de VOORAF-sectie in: kampioen, verrassing, deceptie, topscorer, goals en kaarten.",
    "Groepswinnaar, runner-up en beste nummers 3 worden AUTOMATISCH berekend uit je scores — je hoeft die niet in te vullen.",
    "Knock-out voorspellingen (grijs) vul je pas in na de groepsfase, vóór elke volgende ronde.",
    "Wijzigen van groepsfase-voorspellingen kan niet meer nadat het WK begonnen is.",
    "In principe worden de FIFA WK-regels aangehouden. Bij twijfel beslist de organisator.",
]

PUNTENTELLING = [
    ("Groepswedstrijden:",   "Goede toto (winst/gelijkspel/verlies): 3 pt.  Goede uitslag: +2 pt."),
    ("Knock-outrondes:",     "Goede toto [+ goede uitslag]: R32 5[+3] · R16 7[+4] · KF 9[+5] · HF 11[+6] · F 13[+7] pt.  Uitslag na 90 min telt."),
    ("Doorgaan (top-2):",    "Per goed voorspeld land dat 1e of 2e eindigt in zijn groep: 3 pt."),
    ("Beste nummers 3:",     "Per goed voorspelde nummer 3 die doorgaat: 3 pt  (8 plekken totaal)."),
    ("Kampioen:",            "Verliezend finalist voorspeld: 16 pt.  Kampioen exact: 40 pt."),
    ("Topscorer:",           "3e op topscorerslijst: 9 pt · 2e: 18 pt · 1e: 35 pt.  Exact aantal goals: +8 pt."),
    ("Totaal goals:",        "Exact goed: 25 pt.  Per goal verschil: -1 pt."),
    ("Verrassing:",          "Land buiten top-12 · punten schalen met hoe ver het komt (R16→winnaar) én hoe laag FIFA-gerankt. Zie tabel."),
    ("Deceptie:",            "Land uit top-12 · punten schalen met hoe vroeg het uitvalt (groepsfase=meest) én hoe hoog gerankt. Zie tabel."),
    ("Gele/rode kaarten:",   "Exact goed: 12 pt.  Per kaart verschil: -1 pt.  Geel en rood apart."),
]

RR_PAIRS = [[0,1],[2,3],[0,2],[3,1],[3,0],[1,2]]

def build_group_matches():
    matches = []
    for g, teams in GROUPS.items():
        for n, (a, b) in enumerate(RR_PAIRS, 1):
            matches.append({'id': f'{g}{n}', 'group': g, 'home': teams[a], 'away': teams[b]})
    return matches

GROUP_MATCHES = build_group_matches()
ALL_TEAMS  = [t for teams in GROUPS.values() for t in teams]
OUTSIDERS  = sorted([t for t in ALL_TEAMS if t not in FAVORIETEN], key=lambda l: RANKING.get(l, 99))

def surprise_pts(land, ronde):
    base = {'R16':7,'KF':12,'HF':18,'F':25,'winner':33}.get(ronde, 0)
    pos  = RANKING.get(land, 50)
    return round(base * (pos / 13) ** 0.45)

def deception_pts(land, ronde):
    base = {'KF':6,'R16':11,'R32':16,'groep':24}.get(ronde, 0)
    pos  = RANKING.get(land, 13)
    return round(base * (13 - pos + 3) / 12)

# ─── Stijlen (kleuren exact zoals index.html) ─────────────────────────────────

def fl(hex6):   return PatternFill("solid", fgColor=hex6)
def sd(st, c):  return Side(style=st, color=c)
def bord(st, c): s=sd(st,c); return Border(left=s,right=s,top=s,bottom=s)

FILL_NAVY   = fl("2F6DB5")   # .grouphdr
FILL_INPUT  = fl("BCD2EC")   # .pred  — invulbaar
FILL_HDR    = fl("EEF2F8")   # .colhead / subhdr achtergrond
FILL_GREY   = fl("E0E6EF")   # automatisch berekend / niet invulbaar
FILL_WHITE  = fl("FFFFFF")
FILL_PTBLUE = fl("3B6FB5")   # .pt (punten blauw)
FILL_PTRED  = fl("B22234")   # .pt-red (vooraf-punten rood)
FILL_YELLOW = fl("FFF6D8")   # naam-invulvak
FILL_TOTAL  = fl("FFE14D")   # .total-cell

BORD_GRID   = bord('thin',   'C9D2E0')
BORD_INPUT  = bord('medium', '2A5298')
BORD_GREY   = bord('thin',   'B9C4D6')
BORD_NAAM   = bord('medium', '1A1A1A')

AL_C = Alignment(horizontal='center', vertical='center')
AL_L = Alignment(horizontal='left',   vertical='center')
AL_R = Alignment(horizontal='right',  vertical='center')
AL_W = Alignment(horizontal='left',   vertical='top', wrap_text=True)

def fnt(bold=False, color="15233F", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

# ─── Kolom-layout (15 kolommen, spiegelt website matchCells-structuur) ────────
#
# Website matchCells per wedstrij: cblock + home + cblock + away + pred + res + pt  (7 cellen)
# Twee groepen naast elkaar + spacer = 7 + 1 + 7 = 15 kolommen
#
# Col:  1    2     3    4     5      6      7    8    9    10    11   12    13     14     15
#       clr  home  clr  away  PRED   res    pt  spc  clr  home  clr  away  PRED   res    pt
# px:   16   80    16   80    62     62     32   14   16   80    16   80    62     62     32
# xl:   2    17    2    17    10     10     5    2    2    17    2    17    10     10     5

NCOLS = 15

# Kolommen 7-9 breder zodat verrassing/deceptie-tabelkoppen (Finale, Winnaar, Groep eruit) passen.
# De spacer (8) en kleurblok (9) in de wedstrijdopmaak worden daardoor iets wijder — visueel acceptabel.
COL_W = {1:2, 2:17, 3:2, 4:17, 5:10, 6:10, 7:11, 8:11, 9:11,
         10:17, 11:2, 12:17, 13:10, 14:10, 15:11}

# Kolomnamen voor leesbaarheid
C_CLR1, C_HOME, C_CLR2, C_AWAY, C_PRED, C_RES, C_PT, C_SPC = 1, 2, 3, 4, 5, 6, 7, 8
C_CLR3, C_HOME2, C_CLR4, C_AWAY2, C_PRED2, C_RES2, C_PT2   = 9, 10, 11, 12, 13, 14, 15

# ─── Helpers ──────────────────────────────────────────────────────────────────

def set_col_widths(ws):
    for c, w in COL_W.items():
        ws.column_dimensions[get_column_letter(c)].width = w

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

def mc(ws, row, c1, c2):
    if c1 < c2:
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    return ws.cell(row=row, column=c1)

def sec_hdr(ws, row, text):
    c = mc(ws, row, 1, NCOLS)
    c.value, c.font, c.fill, c.alignment = text, fnt(bold=True, color="FFFFFF", size=11), FILL_NAVY, AL_L
    rh(ws, row, 18)
    return row + 1

def sub_hdr(ws, row, text, color="2A5298", size=11):
    c = mc(ws, row, 1, NCOLS)
    c.value, c.font, c.alignment = text, fnt(bold=True, color=color, size=size), AL_L
    rh(ws, row, 18)
    return row + 1

def empty(ws, row, h=6):
    rh(ws, row, h)
    return row + 1

def inp(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.fill, c.font, c.border, c.alignment = FILL_INPUT, fnt(bold=True, color="0A2A55"), BORD_INPUT, AL_C
    return c

def grey(ws, row, c1, c2=None, text=""):
    c = mc(ws, row, c1, c2 or c1)
    c.value, c.fill, c.font, c.border, c.alignment = text, FILL_GREY, fnt(italic=True, color="888888"), BORD_GREY, AL_C
    return c

def lbl(ws, row, c1, c2, text):
    c = mc(ws, row, c1, c2)
    c.value, c.font, c.alignment = text, fnt(bold=True, color="2A5298"), AL_R
    return c

def dropdown(ws, row, col, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=row, column=col))

def team_clr(ws, row, col, team):
    c = ws.cell(row=row, column=col)
    c.fill = fl(TEAMCOLORS.get(team, "cccccc"))
    c.border = BORD_GRID

def team_name(ws, row, col, name, align=AL_L):
    c = ws.cell(row=row, column=col)
    c.value, c.font, c.border, c.alignment = name, fnt(), BORD_GRID, align

# ─── Sheet builder ────────────────────────────────────────────────────────────

def build_sheet(ws, ls):  # ls = naam van het Landen-sheet
    set_col_widths(ws)
    ws.sheet_view.showGridLines = False
    row = 1

    # ── Naam ─────────────────────────────────────────────────────────────────
    rh(ws, row, 32)
    t = mc(ws, row, 1, 4)
    t.value, t.font, t.alignment = "TEMPETOTO 2026", fnt(bold=True, color="0A1F47", size=16), AL_L

    lbl_c = mc(ws, row, 5, 6)
    lbl_c.value, lbl_c.font, lbl_c.alignment = "Naam:", fnt(bold=True, color="2A5298", size=12), AL_R

    naam = mc(ws, row, 7, 11)
    naam.fill, naam.font, naam.border, naam.alignment = FILL_YELLOW, fnt(bold=True, color="0A1F47", size=13), BORD_NAAM, AL_C

    row = empty(ws, row + 1, 8)

    # ── Vooraf invullen ───────────────────────────────────────────────────────
    all_f = f"={ls}!$A$1:$A${len(ALL_TEAMS)}"
    out_f = f"={ls}!$B$1:$B${len(OUTSIDERS)}"
    fav_f = f"={ls}!$C$1:$C${len(FAVORIETEN)}"

    row = sec_hdr(ws, row, "▶  VOORAF INVULLEN  ·  deadline vóór de eerste wedstrijd")

    prematch = [
        ("Kampioen:",                              all_f),
        ("Verrassing  (outsider, buiten top-12):", out_f),
        ("Deceptie  (favoriet, uit top-12):",      fav_f),
        ("Topscorer  (naam speler):",               None),
        ("↳ Verwacht aantal goals topscorer:",      None),
        ("Totaal goals  (heel WK):",               None),
        ("Gele kaarten  (totaal WK):",             None),
        ("Rode kaarten  (totaal WK):",             None),
    ]
    for label_text, formula in prematch:
        rh(ws, row, 18)
        lbl(ws, row, 1, 4, label_text)
        # Merge over 3 kolommen: alle 4 randen volledig zichtbaar
        c = mc(ws, row, C_PRED, C_PRED + 2)
        c.fill, c.font, c.border, c.alignment = FILL_INPUT, fnt(bold=True, color="0A2A55"), BORD_INPUT, AL_C
        if formula:
            dropdown(ws, row, C_PRED, formula)
        row += 1

    row = empty(ws, row, 8)

    # ── Spelregels ────────────────────────────────────────────────────────────
    row = sub_hdr(ws, row, "Spelregels:")
    for i, regel in enumerate(SPELREGELS, 1):
        rh(ws, row, 40)
        c = mc(ws, row, 1, NCOLS)
        c.value = f"{i}.  {regel}"
        c.font, c.alignment = fnt(color="333333"), AL_W
        row += 1
    row = empty(ws, row, 8)

    # ── Puntentelling ─────────────────────────────────────────────────────────
    row = sub_hdr(ws, row, "Puntentelling:", color="B22234")
    for kop, tekst in PUNTENTELLING:
        # Label-rij
        rh(ws, row, 16)
        c = mc(ws, row, 1, NCOLS)
        c.value, c.font, c.alignment = kop, fnt(bold=True, color="2A5298"), AL_L
        row += 1
        # Tekst-rij, licht ingesprongen
        rh(ws, row, 40)
        c = mc(ws, row, 2, NCOLS)
        c.value, c.font, c.alignment = tekst, fnt(color="444444"), AL_W
        row += 1
    row = empty(ws, row, 8)

    # ── Groepswedstrijden ─────────────────────────────────────────────────────
    row = sec_hdr(ws, row, "▶  GROEPSWEDSTRIJDEN  ·  vul in als '2-1' (thuis–uit, na 90 min)  ·  lichtblauwe vakjes invullen")

    group_keys = list(GROUPS.keys())
    for i in range(0, 12, 2):
        gl, gr = group_keys[i], group_keys[i + 1]
        ml = [m for m in GROUP_MATCHES if m['group'] == gl]
        mr = [m for m in GROUP_MATCHES if m['group'] == gr]

        # Groepskop-rij (Groep X | voorsp. | uitslag | pt. | spacer | Groep Y | ...)
        rh(ws, row, 16)
        for cols, naam in [((C_CLR1, C_AWAY), f"Groep {gl}"), ((C_CLR3, C_AWAY2), f"Groep {gr}")]:
            c = mc(ws, row, cols[0], cols[1])
            c.value, c.font, c.fill, c.alignment = naam, fnt(bold=True, color="2A5298", size=11), FILL_HDR, AL_C
        for col, txt in [(C_PRED,"voorsp."),(C_RES,"uitslag"),(C_PT,"pt."),
                         (C_PRED2,"voorsp."),(C_RES2,"uitslag"),(C_PT2,"pt.")]:
            c = ws.cell(row=row, column=col)
            c.value, c.font, c.fill, c.alignment = txt, fnt(bold=True, color="777777", size=9), FILL_HDR, AL_C
        row += 1

        # 6 wedstrijden
        for k in range(6):
            rh(ws, row, 17)
            for clr_col, nm_col, away_col, pred_col, m, align in [
                (C_CLR1, C_HOME,  C_CLR2, C_PRED,  ml[k], AL_L),
                (C_CLR3, C_HOME2, C_CLR4, C_PRED2, mr[k], AL_L),
            ]:
                team_clr(ws, row, clr_col, m['home'])
                team_name(ws, row, nm_col, m['home'])
                team_clr(ws, row, away_col, m['away'])
                # away team: the column after clr is nm_col+1, after second clr is nm_col+2
                away_nm_col = nm_col + 2
                team_name(ws, row, away_nm_col, m['away'], AL_R)
                inp(ws, row, pred_col)
                # res + pt: grey (niet invulbaar)
                grey(ws, row, pred_col + 1)
                grey(ws, row, pred_col + 2)
            row += 1

        # Groepswinnaar + Runner-up: niet invullen, agent berekent uit scores
        for adv_lbl in ["Groepswinnaar:", "Runner-up:"]:
            rh(ws, row, 17)
            lbl(ws, row, C_CLR1, C_AWAY, adv_lbl)
            grey(ws, row, C_PRED, C_PT, "↳ niet invullen")
            lbl(ws, row, C_CLR3, C_AWAY2, adv_lbl)
            grey(ws, row, C_PRED2, C_PT2, "↳ niet invullen")
            row += 1

        row = empty(ws, row, 6)

    row = empty(ws, row, 8)

    # ── Beste 8 nummers 3 ─────────────────────────────────────────────────────
    row = sec_hdr(ws, row, "▶  BESTE 8 NUMMERS 3  ·  niet invullen — wordt na inlevering door de agent berekend")
    rh(ws, row, 40)
    c = mc(ws, row, 1, NCOLS)
    c.value = ("De 8 beste nummers 3 worden na inlevering van dit formulier automatisch berekend "
               "op basis van de punten en het doelsaldo van de nummer-3-landen in jouw groepsscores. "
               "De Excel zelf rekent dit NIET uit — je hoeft hier niets in te vullen.")
    c.font, c.fill, c.alignment = fnt(italic=True, color="555555"), FILL_GREY, AL_W
    row = empty(ws, row + 1, 8)

    # ── Knock-out (placeholders — later invullen) ─────────────────────────────
    row = sec_hdr(ws, row, "▶  KNOCK-OUT  ·  grijze vakjes — invullen na de groepsfase, vóór elke ronde")

    for ronde in KO_ROUNDS:
        rh(ws, row, 16)
        c = mc(ws, row, 1, NCOLS)
        c.value = f"{ronde['naam']}  ·  toto {ronde['toto']} pt / exact +{ronde['exact']} pt"
        c.font, c.alignment = fnt(bold=True, color="2A5298"), AL_L
        row += 1

        # Kolomkoppen per ronde
        rh(ws, row, 14)
        for col, txt in [(C_PRED,"voorsp."),(C_RES,"uitslag"),(C_PT,"pt.")]:
            c = ws.cell(row=row, column=col)
            c.value, c.font, c.fill, c.alignment = txt, fnt(bold=True, color="777777", size=9), FILL_HDR, AL_C
        row += 1

        brackets = KO_BRACKETS[ronde['key']]
        for home, away in brackets:
            rh(ws, row, 16)
            c = mc(ws, row, C_CLR1, C_AWAY)
            c.value = f"{home}  –  {away}"
            c.font, c.fill, c.border, c.alignment = fnt(italic=True, color="888888"), FILL_GREY, BORD_GREY, AL_L
            grey(ws, row, C_PRED, text="")
            grey(ws, row, C_RES)
            grey(ws, row, C_PT)
            row += 1

        row = empty(ws, row, 6)

    row = empty(ws, row, 8)

    # ── Verrassing-tabel ──────────────────────────────────────────────────────
    row = sec_hdr(ws, row, "▶  VERRASSING — kies een land BUITEN de top-12 · punten schalen met ranking en ronde")

    rh(ws, row, 14)
    c = mc(ws, row, 1, 4)
    c.value, c.font, c.fill, c.alignment = "Land (outsider)", fnt(bold=True,color="777777",size=9), FILL_HDR, AL_C
    for col_i, naam in enumerate(["R16","KF","HF","Finale","Winnaar"], start=5):
        c = ws.cell(row=row, column=col_i)
        c.value, c.font, c.fill, c.alignment = naam, fnt(bold=True,color="777777",size=9), FILL_HDR, AL_C
    row += 1

    for land in OUTSIDERS:
        rh(ws, row, 15)
        c = mc(ws, row, 1, 4)
        c.value, c.font, c.border, c.alignment = land, fnt(), BORD_GRID, AL_L
        for col_i, ronde in enumerate(["R16","KF","HF","F","winner"], start=5):
            c = ws.cell(row=row, column=col_i)
            pts = surprise_pts(land, ronde)
            c.value, c.font, c.fill, c.border, c.alignment = pts, fnt(bold=True,color="FFFFFF"), FILL_PTBLUE, BORD_GRID, AL_C
        row += 1

    row = empty(ws, row, 8)

    # ── Deceptie-tabel ────────────────────────────────────────────────────────
    row = sec_hdr(ws, row, "▶  DECEPTIE — kies een land UIT de top-12 · punten schalen met ranking en uitvalsronde")

    rh(ws, row, 14)
    c = mc(ws, row, 1, 4)
    c.value, c.font, c.fill, c.alignment = "Land (favoriet)", fnt(bold=True,color="777777",size=9), FILL_HDR, AL_C
    for col_i, naam in enumerate(["KF","R16","R32","Groep eruit"], start=5):
        c = ws.cell(row=row, column=col_i)
        c.value, c.font, c.fill, c.alignment = naam, fnt(bold=True,color="777777",size=9), FILL_HDR, AL_C
    row += 1

    for land in FAVORIETEN:
        rh(ws, row, 15)
        c = mc(ws, row, 1, 4)
        c.value, c.font, c.border, c.alignment = land, fnt(), BORD_GRID, AL_L
        for col_i, ronde in enumerate(["KF","R16","R32","groep"], start=5):
            c = ws.cell(row=row, column=col_i)
            pts = deception_pts(land, ronde)
            c.value, c.font, c.fill, c.border, c.alignment = pts, fnt(bold=True,color="FFFFFF"), FILL_PTBLUE, BORD_GRID, AL_C
        row += 1

# ─── Hidden Landen-sheet voor dropdowns ──────────────────────────────────────

def build_landen_sheet(ws):
    ws.sheet_state = "hidden"
    for i, t in enumerate(ALL_TEAMS,  1): ws.cell(i, 1).value = t   # A: alle landen
    for i, t in enumerate(OUTSIDERS,  1): ws.cell(i, 2).value = t   # B: outsiders
    for i, t in enumerate(FAVORIETEN, 1): ws.cell(i, 3).value = t   # C: favorieten

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    landen = wb.active
    landen.title = "Landen"
    build_landen_sheet(landen)

    ws = wb.create_sheet(title="Invulformulier")
    build_sheet(ws, "Landen")

    path = "/home/floris/Tempetoto/tempetoto2026_invulformulier.xlsx"
    wb.save(path)
    print(f"Opgeslagen: {path}")

if __name__ == "__main__":
    main()
